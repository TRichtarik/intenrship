# -*- coding: utf-8 -*-
"""Fill Vykaz_prace.xlsx from timesheet.txt with proposal-aligned descriptions."""
import re
from datetime import date, timedelta
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

MONTHS = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}

DAY_NAMES_SK = ["Po", "Ut", "St", "St", "Pi", "So", "Ne"]
DAY_NAMES_TEMPLATE = ["Po/Mo", "Ut/Tue", "St/Wed", "Ct/Thur", "Pt/Fri"]

HOLIDAYS = {
    date(2026, 4, 3),
    date(2026, 4, 6),
    date(2026, 5, 1),
    date(2026, 5, 8),
}

INTERNSHIP_START = date(2026, 3, 30)
REPORT_START = date(2026, 4, 1)
REPORT_END = date(2026, 6, 30)
TARGET_HOURS = 200
MAX_DAILY_HOURS = 8.0
MAX_WEEKLY_HOURS = 40.0
DESC_TARGET = 130
DESC_MAX = 150

def parse_time(t: str) -> float:
    t = t.strip()
    if not t or t == "00:00":
        return 0.0
    h, m = map(int, t.split(":"))
    return h + m / 60


def parse_date(s: str, default_year: int = 2026) -> date | None:
    s = s.strip()
    m = re.match(r"(\w{3})\s+(\d{1,2})", s)
    if not m:
        return None
    return date(default_year, MONTHS[m.group(1)], int(m.group(2)))


def parse_timesheet(path: str) -> dict[date, list[dict]]:
    with open(path, encoding="utf-8") as f:
        lines = [line.rstrip("\n") for line in f]

    entries: list[dict] = []
    i = 8
    while i < len(lines):
        line = lines[i]
        if not line.strip() or line.strip().startswith("340:"):
            break

        dm = re.match(r"^(\w{3}\s+\d{1,2})\t(.+)$", line)
        if not dm:
            i += 1
            continue

        d = parse_date(dm.group(1))
        rest = dm.group(2)
        parts = [p.strip() for p in rest.split("\t")]
        time_idx = next(
            (idx for idx, part in enumerate(parts) if re.fullmatch(r"\d+:\d{2}", part)),
            None,
        )

        if time_idx is not None:
            project = parts[0]
            task = parts[1] if len(parts) > 1 else ""
            desc = " – ".join(parts[2:time_idx]) if time_idx > 2 else ""
            hours = parse_time(parts[time_idx])
            if d and hours > 0:
                entries.append(
                    {
                        "date": d,
                        "project": project,
                        "task": task,
                        "desc": desc,
                        "hours": hours,
                    }
                )
            i += 1
            continue

        project = rest.strip()
        i += 1
        if i >= len(lines):
            break
        parts = [p.strip() for p in lines[i].split("\t")]
        time_idx = next(
            (idx for idx, part in enumerate(parts) if re.fullmatch(r"\d+:\d{2}", part)),
            None,
        )
        if time_idx is not None:
            task = parts[0]
            desc = " – ".join(parts[1:time_idx]) if time_idx > 1 else ""
            hours = parse_time(parts[time_idx])
            if d and hours > 0:
                entries.append(
                    {
                        "date": d,
                        "project": project,
                        "task": task,
                        "desc": desc,
                        "hours": hours,
                    }
                )
        i += 1

    by_day: dict[date, list[dict]] = defaultdict(list)
    for e in entries:
        by_day[e["date"]].append(e)
    return by_day


def next_workday(d: date) -> date:
    candidate = d + timedelta(days=1)
    while candidate.weekday() >= 5 or candidate in HOLIDAYS:
        candidate += timedelta(days=1)
    return candidate


def previous_workday(d: date) -> date:
    candidate = d - timedelta(days=1)
    while candidate.weekday() >= 5 or candidate in HOLIDAYS:
        candidate -= timedelta(days=1)
    return candidate


def normalize_to_workdays(
    by_day: dict[date, list[dict]],
    start: date,
    end: date,
) -> dict[date, list[dict]]:
    normalized: dict[date, list[dict]] = defaultdict(list)
    overflow: list[dict] = []

    for d in sorted(by_day):
        if d < start or d > end:
            continue
        for entry in by_day[d]:
            if d.weekday() >= 5 or d in HOLIDAYS:
                overflow.append(entry)
            else:
                normalized[d].append(entry)

    for entry in overflow:
        target = next_workday(entry["date"])
        if target > end:
            target = previous_workday(entry["date"])
        if target < start:
            target = start
            while target.weekday() >= 5 or target in HOLIDAYS:
                target += timedelta(days=1)
        normalized[target].append({**entry, "date": target})

    return normalized


def short_project_name(project: str) -> str:
    p = project.strip()
    mapping = {
        "Ecommerce": "Ecommerce",
        "efinoo": "Efinoo",
        "Ultrafast": "Ultrafast Systems",
        "Avalon": "Avalon IT",
        "Inowis": "Inowis",
        "IOTEQ": "IOTEQ",
        "Egne": "Egne",
        "KEMA": "KEMA SK",
        "DODO": "DODO Services",
        "BAUSAD": "BAUSAD",
        "Aviation": "Aviation Group",
        "Inlogic": "Inlogic",
        "Interny": "Interný vývoj",
        "LIANA": "Liana Goliaš",
        "Autospritz": "Autospritz",
    }
    for key, label in mapping.items():
        if key.lower() in p.lower():
            return label
    return "Odoo projekt"


def clean_text(text: str) -> str:
    text = re.sub(r"\s+", " ", text.strip())
    text = text.strip(" –-;,")
    return text


def shorten(text: str, limit: int = 72) -> str:
    text = clean_text(text)
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."


def clean_task_name(task: str) -> str:
    task = re.sub(r"^\[[^\]]+\]\s*", "", task)
    task = re.sub(r"^(DEV|Dev|FE|I|R)\]\s*-?\s*", "", task, flags=re.I)
    task = re.sub(r"^[-–—]\s*", "", task)
    return clean_text(task)


def first_fragment(text: str, limit: int = 48) -> str:
    text = clean_text(text)
    if not text:
        return ""
    for sep in (",", "+", ";", " and ", " – ", " - "):
        if sep in text:
            text = text.split(sep)[0]
    return shorten(text, limit)


TOPIC_LABELS = (
    (r"integrac.*gls|gls sk|gls cz|\bgls\b", "integrácia GLS"),
    (r"integrac.*sps|sps sk|\bsps\b", "integrácia SPS"),
    (r"packeta|dopravc|dpd|raben|vedos", "integrácia dopravcov"),
    (r"gopay", "integrácia GoPay"),
    (r"verzia 1\.3|analyza fe|analyza fe\b", "ecommerce riešenie"),
    (r"ecommerce|eshop|\bfe\b", "ecommerce riešenie"),
    (r"migr.*v19|v19|v18|upgrade", "migrácia Odoo"),
    (r"dfs|dfš|recykl|pravo k platbam", "modul DFS"),
    (r"solidworks|eplan|bom|xls import", "import výrobných dát"),
    (r"audit", "audit riešenia klienta"),
    (r"autodeploy|dev/prod|prostred", "projektové prostredia"),
    (r"web formul", "prepojenie web formulárov"),
    (r"formulár|predplatn|faktúr|faktur", "fakturácia a predplatné"),
    (r"odoo\.sh", "analýza Odoo.sh"),
    (r"inlogic|android|portal", "projekt Inlogic"),
    (r"ultrafast|cmb|reconcil", "projekt Ultrafast Systems"),
    (r"avalon|fintax|sewa|aviation", "projekt Avalon IT"),
    (r"inowis|packeta extend|fintax", "projekt Inowis"),
    (r"kema|eshop cr", "projekt KEMA SK"),
    (r"dodo|sepa|vendor bill", "projekt DODO Services"),
    (r"egne|sensoneo", "projekt Egne"),
    (r"bausad", "projekt BAUSAD"),
    (r"ioteq|shopfloor|eplan|solidworks", "projekt IOTEQ"),
    (r"dokumentac", "projektová dokumentácia"),
    (r"shopfloor|mo lock|email modul", "úpravy ERP procesov"),
)


def looks_english(text: str) -> bool:
    return bool(
        re.search(
            r"\b(the|and|from|with|implement|implementation|create|setup|changes|added|fix|debug|"
            r"deployment|module|study|file|call peto|google|backup|cherry)\b",
            text.lower(),
        )
    )


def detect_topic(task: str, desc: str) -> str:
    text = f"{clean_task_name(task)} {desc}".lower()

    for pattern, label in TOPIC_LABELS:
        if re.search(pattern, text, re.I):
            return label

    task = clean_task_name(task)
    if task and not looks_english(task):
        return shorten(task, 55)

    if desc and not looks_english(first_fragment(desc, 80)):
        return shorten(first_fragment(desc, 55), 55)

    return "implementácia Odoo"


def sanitize_popis(text: str) -> str:
    text = text.replace("–", " ").replace("—", " ")
    text = re.sub(r"-", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def limit_popis(text: str) -> str:
    text = sanitize_popis(text)
    if len(text) <= DESC_MAX:
        return text

    segments = text.split(". ")
    kept: list[str] = []
    for segment in segments:
        candidate = ". ".join(kept + [segment]) if kept else segment
        if len(candidate) <= DESC_MAX:
            kept.append(segment)
        else:
            break

    if kept:
        text = ". ".join(kept)
        if len(text) <= DESC_MAX:
            return text

    if ";" in text:
        head, tail = text.split(":", 1) if ":" in text else ("", text)
        items = [item.strip() for item in tail.split(";")]
        trimmed: list[str] = []
        for item in items:
            prefix = f"{head}: " if head else ""
            candidate = prefix + "; ".join(trimmed + [item])
            if len(candidate) <= DESC_MAX:
                trimmed.append(item)
            else:
                break
        if trimmed:
            return sanitize_popis(
                f"{head}: {'; '.join(trimmed)}" if head else "; ".join(trimmed)
            )

    return sanitize_popis(text[: DESC_MAX - 3].rsplit(" ", 1)[0] + "...")


def reframe_entry(entry: dict) -> str:
    task = clean_task_name(entry["task"])
    desc = clean_text(entry["desc"])
    text = f"{task} {desc}".lower()
    topic = sanitize_popis(detect_topic(entry["task"], desc))

    if re.search(r"\bcall\b", text):
        return sanitize_popis(f"call, koordinácia projektu: {topic}")

    if re.search(r"\b(email|e-mail|followup|follow.up)\b", text):
        return sanitize_popis(f"komunikácia o stave projektu: {topic}")

    if re.search(r"\b(analy|audit|odhad|estimate|study|research|scan)\b", text):
        return sanitize_popis(f"analýza požiadaviek klienta: {topic}")

    if re.search(r"\b(dokument|documentation|dokumentac)\b", text):
        return sanitize_popis(f"príprava dokumentácie: {topic}")

    if re.search(r"\b(test|debug|verify|validation|kontrol)\b", text):
        return sanitize_popis(f"testovanie v Odoo: {topic}")

    if re.search(r"\b(deploy|prod|portainer|cherrypick|backup)\b", text):
        return sanitize_popis(f"koordinácia nasadenia: {topic}")

    if re.search(r"\b(gantt|plan|status|coord)\b", text):
        return sanitize_popis(f"plánovanie projektových úloh: {topic}")

    if re.search(r"\b(implement|modul|parser|wizard|fix|frontend|fe |python|javascript)\b", text):
        return sanitize_popis(f"podpora konfigurácie Odoo: {topic}")

    if re.search(r"\b(integrac|migr)\b", text):
        return sanitize_popis(f"analýza a koordinácia implementácie: {topic}")

    return sanitize_popis(f"sledovanie projektových úloh: {topic}")


def entry_focus(entry: dict) -> str:
    text = " ".join([entry["task"], entry["desc"]]).lower()
    if re.search(
        r"\bcall\b|email|followup|deploy|gantt|plan|status|coord|dokument|presentation",
        text,
    ):
        return "pm"
    if re.search(r"analy|audit|odhad|estimate|study|integrac|migr|research", text):
        return "analysis"
    if re.search(r"test|debug|verify|validation|kontrol", text):
        return "testing"
    if re.search(r"implement|modul|parser|wizard|fix|frontend|python|javascript", text):
        return "dev"
    return "pm"


def build_day_description(day_entries: list[dict], total_hours: float) -> str:
    del total_hours

    by_project: dict[str, list[dict]] = defaultdict(list)
    for entry in day_entries:
        by_project[short_project_name(entry["project"])].append(entry)

    target_mix = ("pm", "pm", "analysis", "analysis", "testing", "dev")
    segments: list[str] = []

    for project, entries in list(by_project.items())[:4]:
        buckets: dict[str, list[str]] = defaultdict(list)
        seen: set[str] = set()

        for entry in sorted(entries, key=lambda item: item["hours"], reverse=True):
            phrase = reframe_entry(entry)
            focus = entry_focus(entry)
            key = phrase.lower()
            if key in seen:
                continue
            seen.add(key)
            buckets[focus].append(phrase)

        selected: list[str] = []
        used: set[str] = set()

        for focus in target_mix:
            for phrase in buckets.get(focus, []):
                if phrase not in used:
                    selected.append(phrase)
                    used.add(phrase)
                    break
            if len(selected) >= 2:
                break

        for focus in ("pm", "analysis", "testing", "dev"):
            for phrase in buckets.get(focus, []):
                if phrase not in used and len(selected) < 2:
                    selected.append(phrase)
                    used.add(phrase)

        if not selected:
            selected = ["koordinácia a analýza projektových úloh: implementácia Odoo"]

        segment = f"{project}: {'; '.join(selected[:2])}"
        candidate = limit_popis(". ".join(segments + [segment]))
        if segments and len(candidate) > DESC_MAX:
            break
        segments.append(segment)
        if len(". ".join(segments)) >= DESC_TARGET - 20:
            break

    return limit_popis(". ".join(segments))


def date_to_row(d: date) -> int | None:
    if d.weekday() >= 5 or d in HOLIDAYS:
        return None
    if d < INTERNSHIP_START:
        return None
    offset = (d - INTERNSHIP_START).days
    week_idx = offset // 7
    day_idx = offset % 7
    if day_idx >= 5:
        return None
    return 6 + week_idx * 5 + day_idx


def sk_plural(n: int, one: str, few: str, many: str) -> str:
    if n == 1:
        return one
    if 2 <= n <= 4:
        return few
    return many


def round_to_half_hour(hours: float) -> float:
    return round(hours * 2) / 2


def week_start(d: date) -> date:
    return d - timedelta(days=d.weekday())


def format_hours(hours: float) -> float:
    return round_to_half_hour(hours)


def format_hours_display(hours: float) -> str:
    value = round_to_half_hour(hours)
    if value.is_integer():
        return str(int(value))
    return str(value).replace(".", ",")


def scale_hours_to_target(day_hours: list[float], target: float) -> list[float]:
    raw_total = sum(day_hours)
    if raw_total <= 0:
        return [0.0] * len(day_hours)

    target_halves = int(round(target * 2))
    fractional = [hours * target_halves / raw_total for hours in day_hours]
    scaled_halves = [round(value) for value in fractional]
    diff = target_halves - sum(scaled_halves)

    if diff > 0:
        order = sorted(
            range(len(day_hours)),
            key=lambda index: fractional[index] - scaled_halves[index],
            reverse=True,
        )
        for index in order:
            if day_hours[index] <= 0 or diff <= 0:
                continue
            scaled_halves[index] += 1
            diff -= 1
    elif diff < 0:
        order = sorted(
            range(len(day_hours)),
            key=lambda index: fractional[index] - scaled_halves[index],
        )
        for index in order:
            if scaled_halves[index] <= 0 or diff >= 0:
                continue
            scaled_halves[index] -= 1
            diff += 1

    return [half / 2 for half in scaled_halves]


def week_indices(pending: list[tuple[date, int, list[dict], float]]) -> dict[date, list[int]]:
    grouped: dict[date, list[int]] = defaultdict(list)
    for index, (day, _, _, _) in enumerate(pending):
        grouped[week_start(day)].append(index)
    return grouped


def week_total(hours: list[float], indices: list[int]) -> float:
    return sum(hours[index] for index in indices)


def day_capacity(
    hours: list[float],
    pending: list[tuple[date, int, list[dict], float]],
    index: int,
    weeks: dict[date, list[int]],
) -> float:
    day_hours = hours[index]
    daily_room = max(0.0, MAX_DAILY_HOURS - day_hours)
    week_room = max(
        0.0,
        MAX_WEEKLY_HOURS - week_total(hours, weeks[week_start(pending[index][0])]),
    )
    return min(daily_room, week_room)


def apply_hour_limits(
    pending: list[tuple[date, int, list[dict], float]],
    scaled_hours: list[float],
    target: float = TARGET_HOURS,
) -> list[float]:
    hours = [min(round_to_half_hour(value), MAX_DAILY_HOURS) for value in scaled_hours]
    weeks = week_indices(pending)

    for indices in weeks.values():
        overflow = round_to_half_hour(week_total(hours, indices) - MAX_WEEKLY_HOURS)
        while overflow > 0:
            reducible = [
                index
                for index in indices
                if hours[index] >= 0.5 and week_total(hours, indices) > MAX_WEEKLY_HOURS
            ]
            if not reducible:
                break
            index = max(reducible, key=lambda item: hours[item])
            hours[index] = round_to_half_hour(hours[index] - 0.5)
            overflow = round_to_half_hour(week_total(hours, indices) - MAX_WEEKLY_HOURS)

    target_halves = int(round(target * 2))
    while round(sum(hours) * 2) < target_halves:
        candidates = [
            index
            for index in range(len(hours))
            if day_capacity(hours, pending, index, weeks) >= 0.5
        ]
        if not candidates:
            break
        index = max(candidates, key=lambda item: pending[item][3])
        hours[index] = round_to_half_hour(hours[index] + 0.5)

    return hours


def main() -> None:
    base = Path(__file__).resolve().parent
    timesheet_path = base / "timesheet.txt"
    xlsx_path = base / "Vykaz_prace.xlsx"

    by_day = parse_timesheet(timesheet_path)
    by_day = normalize_to_workdays(by_day, REPORT_START, REPORT_END)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["List1"]

    ws["B1"] = "Tomáš Richtárik"
    ws["B2"] = "514392"

    # Clear previous entries in the date range rows
    for row in range(6, 271):
        ws.cell(row=row, column=2, value=None)
        ws.cell(row=row, column=3, value=None)
        ws.cell(row=row, column=4, value=None)

    pending: list[tuple[date, int, list[dict], float]] = []
    d = REPORT_START
    while d <= REPORT_END:
        if d.weekday() >= 5 or d in HOLIDAYS:
            d += timedelta(days=1)
            continue

        row = date_to_row(d)
        day_hours = sum(e["hours"] for e in by_day.get(d, []))
        if row and day_hours > 0:
            pending.append((d, row, by_day[d], day_hours))
        d += timedelta(days=1)

    scaled_hours = apply_hour_limits(
        pending,
        scale_hours_to_target([item[3] for item in pending], TARGET_HOURS),
    )

    total = 0.0
    filled_days = 0
    for (d, row, entries, _), day_hours in zip(pending, scaled_hours):
        if day_hours <= 0:
            continue
        desc = build_day_description(entries, day_hours)
        ws.cell(row=row, column=2, value=d)
        ws.cell(row=row, column=3, value=format_hours(day_hours))
        cell_d = ws.cell(row=row, column=4, value=desc)
        cell_d.alignment = Alignment(wrap_text=True, vertical="top")
        total += format_hours(day_hours)
        filled_days += 1

    ws["C3"] = format_hours(total)

    wb.save(xlsx_path)
    print(
        f"Filled {filled_days} work days, total {format_hours_display(total)} h "
        f"(max {int(MAX_DAILY_HOURS)} h/deň, max {int(MAX_WEEKLY_HOURS)} h/týždeň)"
    )
    print(f"Saved to {xlsx_path}")


if __name__ == "__main__":
    main()
